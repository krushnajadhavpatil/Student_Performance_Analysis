import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── DATA LOAD ──────────────────────────────────────────
df = pd.read_csv('data/student_data.csv')
df['Total_Marks'] = df['Internal_Marks'] + df['External_Marks']
df['Status'] = df['External_Marks'].apply(lambda x: 'Pass' if x >= 32 else 'Fail')
df['Backlog'] = df['Status'].apply(lambda x: 1 if x == 'Fail' else 0)

# ── STYLES ─────────────────────────────────────────────
def hdr(ws, row, col, text, bg, txt_color="FFFFFF"):
    c = ws.cell(row=row, column=col, value=text)
    c.font = Font(bold=True, color=txt_color, name="Arial", size=10)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

def cell(ws, row, col, value, bg="FFFFFF", bold=False, color="000000"):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=color, name="Arial", size=10)
    c.fill = PatternFill("solid", fgColor=bg)
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

wb = Workbook()

# ══════════════════════════════════════════════════════
# SHEET 1 — RAW DATA
# ══════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Raw Data"

cols = list(df.columns)
for i, h in enumerate(cols, 1):
    hdr(ws1, 1, i, h.replace("_", " ").title(), "1F4E79")

for r, row in df.iterrows():
    bg = "D6E4F0" if r % 2 == 0 else "FFFFFF"
    for c_i, col in enumerate(cols, 1):
        val = row[col]
        if col == "Status":
            color = "375623" if val == "Pass" else "9C0006"
            bg2   = "C6EFCE" if val == "Pass" else "FFC7CE"
            cell(ws1, r+2, c_i, val, bg2, True, color)
        else:
            cell(ws1, r+2, c_i, val, bg)

widths = [10,16,12,10,20,22,15,17,9,13,8]
for i, w in enumerate(widths, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = "A2"
ws1.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(df)+1}"

print("Sheet 1 - Raw Data done!")

# ══════════════════════════════════════════════════════
# SHEET 2 — SUMMARY DASHBOARD
# ══════════════════════════════════════════════════════
ws2 = wb.create_sheet("Summary Dashboard")
ws2.sheet_view.showGridLines = False

# -- KPI Cards --
ws2.merge_cells("B2:D3")
ws2["B2"] = f"{df['Student_ID'].nunique()}"
ws2["B2"].font = Font(bold=True, size=22, color="1F4E79", name="Arial")
ws2["B2"].alignment = Alignment(horizontal="center", vertical="center")
ws2["B2"].fill = PatternFill("solid", fgColor="D6E4F0")
ws2.merge_cells("B4:D4")
ws2["B4"] = "Total Students"
ws2["B4"].font = Font(bold=True, color="1F4E79", name="Arial", size=10)
ws2["B4"].alignment = Alignment(horizontal="center")

pass_rate = round((df["Status"] == "Pass").mean() * 100, 1)
ws2.merge_cells("F2:H3")
ws2["F2"] = f"{pass_rate}%"
ws2["F2"].font = Font(bold=True, size=22, color="375623", name="Arial")
ws2["F2"].alignment = Alignment(horizontal="center", vertical="center")
ws2["F2"].fill = PatternFill("solid", fgColor="C6EFCE")
ws2.merge_cells("F4:H4")
ws2["F4"] = "Pass Rate"
ws2["F4"].font = Font(bold=True, color="375623", name="Arial", size=10)
ws2["F4"].alignment = Alignment(horizontal="center")

backlogs = int(df["Backlog"].sum())
ws2.merge_cells("J2:L3")
ws2["J2"] = f"{backlogs}"
ws2["J2"].font = Font(bold=True, size=22, color="9C0006", name="Arial")
ws2["J2"].alignment = Alignment(horizontal="center", vertical="center")
ws2["J2"].fill = PatternFill("solid", fgColor="FFC7CE")
ws2.merge_cells("J4:L4")
ws2["J4"] = "Total Backlogs"
ws2["J4"].font = Font(bold=True, color="9C0006", name="Arial", size=10)
ws2["J4"].alignment = Alignment(horizontal="center")

avg_att = round(df["Attendance_Percentage"].mean(), 1)
ws2.merge_cells("N2:P3")
ws2["N2"] = f"{avg_att}%"
ws2["N2"].font = Font(bold=True, size=22, color="7B3F00", name="Arial")
ws2["N2"].alignment = Alignment(horizontal="center", vertical="center")
ws2["N2"].fill = PatternFill("solid", fgColor="FAEEDA")
ws2.merge_cells("N4:P4")
ws2["N4"] = "Avg Attendance"
ws2["N4"].font = Font(bold=True, color="7B3F00", name="Arial", size=10)
ws2["N4"].alignment = Alignment(horizontal="center")

# -- Subject Table --
hdr(ws2, 6, 2, "Subject",          "2E75B6")
hdr(ws2, 6, 3, "Avg Attendance %", "2E75B6")
hdr(ws2, 6, 4, "Avg Total Marks",  "2E75B6")
hdr(ws2, 6, 5, "Pass Rate %",      "2E75B6")
hdr(ws2, 6, 6, "Backlogs",         "2E75B6")

sub = df.groupby("Subject").agg(
    Att=("Attendance_Percentage","mean"),
    Marks=("Total_Marks","mean"),
    Backlogs=("Backlog","sum"),
    Total=("Status","count"),
    Pass=("Status", lambda x: (x=="Pass").sum())
).reset_index()
sub["PR"] = (sub["Pass"]/sub["Total"]*100).round(1)

for i, row in sub.iterrows():
    bg = "D6E4F0" if i%2==0 else "FFFFFF"
    cell(ws2, 7+i, 2, row["Subject"],        bg)
    cell(ws2, 7+i, 3, f'{row["Att"]:.1f}%',  bg)
    cell(ws2, 7+i, 4, f'{row["Marks"]:.1f}', bg)
    cell(ws2, 7+i, 5, f'{row["PR"]:.1f}%',   bg)
    cell(ws2, 7+i, 6, int(row["Backlogs"]),   bg)

# -- Department Table --
hdr(ws2, 6, 9,  "Department",      "375623")
hdr(ws2, 6, 10, "Avg Attendance%", "375623")
hdr(ws2, 6, 11, "Avg Marks",       "375623")
hdr(ws2, 6, 12, "Students",        "375623")
hdr(ws2, 6, 13, "Backlogs",        "375623")

dept = df.groupby("Department").agg(
    Att=("Attendance_Percentage","mean"),
    Marks=("Total_Marks","mean"),
    Students=("Student_ID","nunique"),
    Backlogs=("Backlog","sum")
).reset_index()

for i, row in dept.iterrows():
    bg = "D6E4F0" if i%2==0 else "FFFFFF"
    cell(ws2, 7+i, 9,  row["Department"],      bg)
    cell(ws2, 7+i, 10, f'{row["Att"]:.1f}%',   bg)
    cell(ws2, 7+i, 11, f'{row["Marks"]:.1f}',  bg)
    cell(ws2, 7+i, 12, int(row["Students"]),    bg)
    cell(ws2, 7+i, 13, int(row["Backlogs"]),    bg)

for col in range(2, 17):
    ws2.column_dimensions[get_column_letter(col)].width = 17

print("Sheet 2 - Summary done!")

# ══════════════════════════════════════════════════════
# SHEET 3 — AT-RISK STUDENTS  (Legend added)
# ══════════════════════════════════════════════════════
ws3 = wb.create_sheet("At-Risk Students")
ws3.sheet_view.showGridLines = False

# Title
ws3.merge_cells("A1:G1")
ws3["A1"] = "AT-RISK STUDENTS — Attendance Below 65%"
ws3["A1"].font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
ws3["A1"].fill = PatternFill("solid", fgColor="C00000")
ws3["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[1].height = 28

# Headers
risk_hdrs = ["Student ID","Name","Department","Semester",
             "Avg Attendance","Total Backlogs","Avg Marks"]
for i, h in enumerate(risk_hdrs, 1):
    hdr(ws3, 2, i, h, "FF0000")

# Data
at_risk = df[df["Attendance_Percentage"] < 65].groupby(
    ["Student_ID","Name","Department","Semester"]
).agg(
    Att=("Attendance_Percentage","mean"),
    Backlogs=("Backlog","sum"),
    Marks=("Total_Marks","mean")
).reset_index().sort_values("Att")

at_risk = at_risk.reset_index(drop=True)

for i, row in at_risk.iterrows():
    danger = row["Att"] < 40
    bg  = "FFE0E0" if danger else "FFF2CC"
    txt = "9C0006" if danger else "000000"
    vals = [row["Student_ID"], row["Name"], row["Department"],
            int(row["Semester"]), f'{row["Att"]:.1f}%',
            int(row["Backlogs"]), f'{row["Marks"]:.1f}']
    for j, v in enumerate(vals, 1):
        cell(ws3, 3+i, j, v, bg, danger, txt)

risk_w = [12,18,13,10,16,15,12]
for i, w in enumerate(risk_w, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# ── LEGEND (neeche data ke baad) ──────────────────────
legend_row = 3 + len(at_risk) + 2   # 2 rows gap ke baad

# Legend heading
ws3.merge_cells(
    start_row=legend_row, start_column=1,
    end_row=legend_row,   end_column=7
)
c = ws3.cell(row=legend_row, column=1, value="LEGEND")
c.font    = Font(bold=True, color="FFFFFF", name="Arial", size=10)
c.fill    = PatternFill("solid", fgColor="404040")
c.alignment = Alignment(horizontal="center", vertical="center")
ws3.row_dimensions[legend_row].height = 20

# Legend row 1 — Critical
leg1 = legend_row + 1
ws3.merge_cells(
    start_row=leg1, start_column=1,
    end_row=leg1,   end_column=2
)
c1 = ws3.cell(row=leg1, column=1, value="Critical Student")
c1.font      = Font(bold=True, color="9C0006", name="Arial", size=10)
c1.fill      = PatternFill("solid", fgColor="FFE0E0")
c1.alignment = Alignment(horizontal="center", vertical="center")
c1.border    = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)
ws3.merge_cells(
    start_row=leg1, start_column=3,
    end_row=leg1,   end_column=7
)
c1d = ws3.cell(row=leg1, column=3,
               value="Attendance < 40% — Immediate action needed!")
c1d.font      = Font(color="9C0006", name="Arial", size=10)
c1d.fill      = PatternFill("solid", fgColor="FFE0E0")
c1d.alignment = Alignment(horizontal="left", vertical="center")
c1d.border    = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)

# Legend row 2 — At-Risk
leg2 = legend_row + 2
ws3.merge_cells(
    start_row=leg2, start_column=1,
    end_row=leg2,   end_column=2
)
c2 = ws3.cell(row=leg2, column=1, value="At-Risk Student")
c2.font      = Font(bold=True, color="7B3F00", name="Arial", size=10)
c2.fill      = PatternFill("solid", fgColor="FFF2CC")
c2.alignment = Alignment(horizontal="center", vertical="center")
c2.border    = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)
ws3.merge_cells(
    start_row=leg2, start_column=3,
    end_row=leg2,   end_column=7
)
c2d = ws3.cell(row=leg2, column=3,
               value="Attendance 40–65% — Monitor closely")
c2d.font      = Font(color="7B3F00", name="Arial", size=10)
c2d.fill      = PatternFill("solid", fgColor="FFF2CC")
c2d.alignment = Alignment(horizontal="left", vertical="center")
c2d.border    = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)

print("Sheet 3 - At-Risk + Legend done!")

# ══════════════════════════════════════════════════════
# SHEET 4 — SEMESTER SUMMARY  (Semester + Att Category)
# ══════════════════════════════════════════════════════
ws4 = wb.create_sheet("Semester Summary")
ws4.sheet_view.showGridLines = False

# ── Part A: Semester-wise Performance ─────────────────
ws4.merge_cells("A1:D1")
ws4["A1"] = "Semester-wise Performance Analysis"
ws4["A1"].font      = Font(bold=True, color="FFFFFF", name="Arial", size=12)
ws4["A1"].fill      = PatternFill("solid", fgColor="375623")
ws4["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[1].height = 28

sem_hdrs = ["Semester","Avg Attendance %","Avg Total Marks","Total Backlogs"]
for i, h in enumerate(sem_hdrs, 1):
    hdr(ws4, 2, i, h, "70AD47")

sem = df.groupby("Semester").agg(
    Att=("Attendance_Percentage","mean"),
    Marks=("Total_Marks","mean"),
    Backlogs=("Backlog","sum")
).reset_index()

for i, row in sem.reset_index(drop=True).iterrows():
    bg = "EAF3DE" if i % 2 == 0 else "FFFFFF"
    cell(ws4, 3+i, 1, f"Semester {int(row['Semester'])}", bg)
    cell(ws4, 3+i, 2, f"{row['Att']:.1f}%",              bg)
    cell(ws4, 3+i, 3, f"{row['Marks']:.1f}",             bg)
    cell(ws4, 3+i, 4, int(row['Backlogs']),               bg)

for i in range(1, 5):
    ws4.column_dimensions[get_column_letter(i)].width = 22

# ── Part B: Attendance Category vs Pass Rate ──────────
# 3 rows gap ke baad start karo
part_b_start = 3 + len(sem) + 3

ws4.merge_cells(
    start_row=part_b_start, start_column=1,
    end_row=part_b_start,   end_column=5
)
tb = ws4.cell(row=part_b_start, column=1,
              value="Attendance Category vs Pass Rate")
tb.font      = Font(bold=True, color="FFFFFF", name="Arial", size=12)
tb.fill      = PatternFill("solid", fgColor="2E75B6")
tb.alignment = Alignment(horizontal="center", vertical="center")
ws4.row_dimensions[part_b_start].height = 28

# Headers
cat_hdrs = ["Attendance Category","Total Students",
            "Pass Rate %","Backlog Count","Avg Marks"]
for i, h in enumerate(cat_hdrs, 1):
    hdr(ws4, part_b_start+1, i, h, "4472C4")

# Attendance category column banao
def att_cat(x):
    if x >= 85:   return "High (>=85%)"
    elif x >= 65: return "Medium (65-84%)"
    else:         return "Low (<65%)"

df["Att_Cat"] = df["Attendance_Percentage"].apply(att_cat)

cat_g = df.groupby("Att_Cat").agg(
    Total=("Student_ID","count"),
    Passes=("Status", lambda x: (x=="Pass").sum()),
    Backlogs=("Backlog","sum"),
    Marks=("Total_Marks","mean")
).reset_index()
cat_g["PassRate"] = (cat_g["Passes"] / cat_g["Total"] * 100).round(1)

# Order: High, Medium, Low
order = {"High (>=85%)": 0, "Medium (65-84%)": 1, "Low (<65%)": 2}
cat_g["ord"] = cat_g["Att_Cat"].map(order)
cat_g = cat_g.sort_values("ord").reset_index(drop=True)

# Color per category
cat_colors = ["C6EFCE", "FFEB9C", "FFC7CE"]
cat_txt    = ["375623", "7B3F00", "9C0006"]

for i, row in cat_g.iterrows():
    bg  = cat_colors[i]
    txt = cat_txt[i]
    data_row = part_b_start + 2 + i
    vals = [row["Att_Cat"], int(row["Total"]),
            f'{row["PassRate"]:.1f}%',
            int(row["Backlogs"]),
            f'{row["Marks"]:.1f}']
    for j, v in enumerate(vals, 1):
        c = ws4.cell(row=data_row, column=j, value=v)
        c.font      = Font(bold=True, color=txt, name="Arial", size=10)
        c.fill      = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border    = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'),  bottom=Side(style='thin')
        )

# Column E ko bhi set karo
ws4.column_dimensions["E"].width = 16

print("Sheet 4 - Semester Summary + Att Category done!")

# ── SAVE ───────────────────────────────────────────────
wb.save("Attendance_Dashboard.xlsx")
print("\n✅ Poori Excel file ready hai — 4 sheets ke saath!")